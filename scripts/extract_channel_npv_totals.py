"""
Extract channel-level NPV totals from weekly Excel reporting files.

The script scans weekly Excel workbooks, keeps country total rows only,
extracts NPV metrics, and exports a clean dataset with an issues log.
"""

import os
import re
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook


ROOT = Path(r"PATH_TO_SOURCE_FOLDER")

SOURCE_FOLDERS = {
    "brand_a": ROOT / r"path\to\brand_a",
    "brand_b": ROOT / r"path\to\brand_b",
    "brand_c": ROOT / r"path\to\brand_c",
}

OUTPUT_PATH = Path(r"PATH_TO_OUTPUT_FOLDER\channel_npv_extract.xlsx")

WEEKS_NEEDED = set(range(1, 18))
SHEET_NAME = "Info - Split"
CHANNEL_NAME = "Retail Channel"


def detect_week(filename: str):
    match = re.search(r"week\s*(\d+)", filename, re.I)
    return int(match.group(1)) if match else None


def financial_week(week_num: int):
    return f"W2026{week_num:02d}"


def should_skip_file(path: Path):
    name = path.name.lower()

    if path.name.startswith("~$"):
        return True

    if not name.endswith((".xlsx", ".xlsm")):
        return True

    if "std" in name:
        return True

    week_num = detect_week(path.name)

    return week_num not in WEEKS_NEEDED


def version_number(path: Path):
    match = re.search(r"\bv(\d+)\b", path.stem.lower())
    return int(match.group(1)) if match else 0


def choose_latest_versions(files):
    grouped = {}

    for file_path in files:
        week_num = detect_week(file_path.name)

        if week_num is not None:
            grouped.setdefault(week_num, []).append(file_path)

    chosen = []

    for group in grouped.values():
        group_sorted = sorted(
            group,
            key=lambda file_path: (version_number(file_path), file_path.stat().st_mtime),
        )
        chosen.append(group_sorted[-1])

    return sorted(chosen, key=lambda file_path: detect_week(file_path.name) or 0)


def clean_text(value):
    return "" if value is None else str(value).strip()


def clean_country(value):
    text = clean_text(value)
    text = re.sub(r"^retail channel\s*-\s*", "", text, flags=re.I)
    text = re.sub(r"\btotal\b", "", text, flags=re.I)
    text = re.sub(r"\s+", " ", text).strip()
    return text.upper()


def to_number(value):
    if value is None:
        return None

    text = str(value).strip()

    if text in ["", "-", "#N/A", "N/A", "nan", "None"]:
        return None

    text = text.replace("£", "").replace(",", "").replace("%", "").strip()

    try:
        return float(text)
    except ValueError:
        return None


def get_sheet_case_insensitive(workbook, wanted_name):
    wanted_clean = wanted_name.lower().replace(" ", "").replace("-", "")

    for sheet in workbook.sheetnames:
        sheet_clean = sheet.lower().replace(" ", "").replace("-", "")

        if sheet_clean == wanted_clean:
            return sheet

    return None


def extract_file(path: Path, brand: str):
    rows = []
    issues = []

    week_num = detect_week(path.name)

    if week_num is None:
        issues.append({"file": str(path), "issue": "Could not detect week from filename"})
        return rows, issues

    try:
        workbook = load_workbook(path, read_only=True, data_only=True)
    except Exception as error:
        issues.append({"file": str(path), "issue": f"Could not open workbook: {error}"})
        return rows, issues

    sheet_name = get_sheet_case_insensitive(workbook, SHEET_NAME)

    if sheet_name is None:
        issues.append({
            "file": str(path),
            "issue": f"{SHEET_NAME} sheet not found. Sheets found: {workbook.sheetnames}",
        })
        workbook.close()
        return rows, issues

    worksheet = workbook[sheet_name]

    for row_num, row in enumerate(worksheet.iter_rows(values_only=True), start=1):
        values = list(row)

        total_label = clean_text(values[1] if len(values) > 1 else "")

        if "total" not in total_label.lower():
            continue

        if total_label.strip().lower() == "total":
            continue

        country = clean_country(total_label)

        if country == "":
            continue

        row_text = " ".join(clean_text(value).lower() for value in values[:8])

        if "channel" in row_text or "country" in row_text or "level" in row_text:
            continue

        npv = to_number(values[39] if len(values) > 39 else None)
        npv_per_1 = to_number(values[40] if len(values) > 40 else None)

        if npv is None and npv_per_1 is None:
            continue

        rows.append({
            "financial_week": financial_week(week_num),
            "country": country,
            "channel": CHANNEL_NAME,
            "brand": brand,
            "npv": npv,
            "npv_per_1": npv_per_1,
            "source_file": path.name,
            "sheet_used": sheet_name,
            "source_row": row_num,
        })

    workbook.close()
    return rows, issues


def main():
    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)

    all_rows = []
    all_issues = []

    for brand, folder in SOURCE_FOLDERS.items():
        print(f"\nChecking {CHANNEL_NAME} / {brand}")

        if not folder.exists():
            all_issues.append({
                "brand": brand,
                "file": str(folder),
                "issue": "Folder does not exist",
            })
            continue

        files = []

        for dirpath, _, filenames in os.walk(folder):
            for filename in filenames:
                file_path = Path(dirpath) / filename

                if not should_skip_file(file_path):
                    files.append(file_path)

        files = choose_latest_versions(files)

        print(f"Files selected: {len(files)}")

        for index, file_path in enumerate(files, start=1):
            print(f"[{index}/{len(files)}] {file_path.name}")

            rows, issues = extract_file(file_path, brand)
            all_rows.extend(rows)

            for issue in issues:
                issue["brand"] = brand
                all_issues.append(issue)

    data = pd.DataFrame(all_rows)
    issues = pd.DataFrame(all_issues)

    with pd.ExcelWriter(OUTPUT_PATH, engine="openpyxl") as writer:
        data.to_excel(writer, sheet_name="cleaned_data", index=False)
        issues.to_excel(writer, sheet_name="issues", index=False)

    print("\nDone")
    print(f"Saved to: {OUTPUT_PATH}")
    print(f"Rows extracted: {len(data)}")
    print(f"Issues: {len(issues)}")


if __name__ == "__main__":
    main()
