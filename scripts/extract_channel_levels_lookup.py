from pathlib import Path
import re

import pandas as pd
from openpyxl import load_workbook


ROOT_FOLDER = Path("path/to/source_folder")
OUTPUT_FILE = Path("path/to/output_file.xlsx")

PERIODS_TO_USE = ["P1", "P2", "P3", "P4", "P5"]

MAIN_SHEET_NAME = "Main"
LOOKUP_SHEET_NAME = "Lookup"

MAIN_COLUMNS = {
    "category": "B",
    "level_1": "C",
    "level_2": "D",
    "metric_1": "G",
}

LOOKUP_COLUMNS = {
    "category": "B",
    "level_1": "C",
    "level_2": "D",
    "metric_2": "AL",
    "metric_3": "AM",
}


def clean_key_value(value):
    if value is None:
        return ""

    return str(value).strip().lower()


def make_lookup_key(category, level_1, level_2):
    return (
        clean_key_value(category),
        clean_key_value(level_1),
        clean_key_value(level_2),
    )


def extract_period(file_path):
    for part in file_path.parts:
        if part.upper().startswith("P") and part[1:].isdigit():
            return part.upper()

    return None


def extract_week(file_path):
    match = re.search(
        r"(?:WK|WEEK|WEEKLY)\s*0?(\d{1,2})",
        file_path.name,
        re.IGNORECASE,
    )

    if match:
        return int(match.group(1))

    return None


def is_weekly_file(file_path):
    full_path = str(file_path).lower()
    file_name = file_path.name.lower()

    if file_path.suffix.lower() not in [".xlsx", ".xlsm"]:
        return False

    if file_name.startswith("~$"):
        return False

    if "std" in full_path:
        return False

    if "archive" in full_path:
        return False

    return "wk" in file_name or "week" in file_name or "weekly" in file_name


def get_version_number(file_path):
    file_name = file_path.name.lower()

    match = re.search(r"v(\d+)", file_name)

    if match:
        return int(match.group(1))

    return 1


def get_weekly_files():
    files = []

    for period in PERIODS_TO_USE:
        period_folder = ROOT_FOLDER / period

        if period_folder.exists():
            for file in period_folder.rglob("*"):
                if file.is_file() and is_weekly_file(file):
                    files.append(file)
        else:
            print("Missing period folder:", period_folder)

    for file in ROOT_FOLDER.iterdir():
        if file.is_file() and is_weekly_file(file):
            files.append(file)

    latest_files_by_week = {}

    for file in files:
        week = extract_week(file)

        if week is None:
            continue

        current_file = latest_files_by_week.get(week)

        if current_file is None:
            latest_files_by_week[week] = file
            continue

        if get_version_number(file) > get_version_number(current_file):
            latest_files_by_week[week] = file

    return [
        latest_files_by_week[week]
        for week in sorted(latest_files_by_week)
    ]


def build_lookup(worksheet):
    lookup = {}

    for row in range(1, worksheet.max_row + 1):
        category = worksheet[f"{LOOKUP_COLUMNS['category']}{row}"].value
        level_1 = worksheet[f"{LOOKUP_COLUMNS['level_1']}{row}"].value
        level_2 = worksheet[f"{LOOKUP_COLUMNS['level_2']}{row}"].value
        metric_2 = worksheet[f"{LOOKUP_COLUMNS['metric_2']}{row}"].value
        metric_3 = worksheet[f"{LOOKUP_COLUMNS['metric_3']}{row}"].value

        key = make_lookup_key(category, level_1, level_2)

        if not any(key):
            continue

        if any("total" in item for item in key):
            continue

        lookup[key] = {
            "Metric 2": metric_2,
            "Metric 3": metric_3,
        }

    return lookup


def extract_file(file_path):
    rows = []
    log_entry = {
        "Period": extract_period(file_path),
        "Week": extract_week(file_path),
        "File": file_path.name,
        "Status": "",
        "Reason": "",
        "Rows Extracted": 0,
        "Lookup Misses": 0,
    }

    try:
        workbook = load_workbook(file_path, data_only=True, read_only=True)

        if MAIN_SHEET_NAME not in workbook.sheetnames:
            log_entry["Status"] = "Skipped"
            log_entry["Reason"] = f"Missing sheet: {MAIN_SHEET_NAME}"
            workbook.close()
            return rows, log_entry

        if LOOKUP_SHEET_NAME not in workbook.sheetnames:
            log_entry["Status"] = "Skipped"
            log_entry["Reason"] = f"Missing sheet: {LOOKUP_SHEET_NAME}"
            workbook.close()
            return rows, log_entry

        main_worksheet = workbook[MAIN_SHEET_NAME]
        lookup_worksheet = workbook[LOOKUP_SHEET_NAME]

        lookup = build_lookup(lookup_worksheet)

        for row in range(1, main_worksheet.max_row + 1):
            category = main_worksheet[f"{MAIN_COLUMNS['category']}{row}"].value
            level_1 = main_worksheet[f"{MAIN_COLUMNS['level_1']}{row}"].value
            level_2 = main_worksheet[f"{MAIN_COLUMNS['level_2']}{row}"].value
            metric_1 = main_worksheet[f"{MAIN_COLUMNS['metric_1']}{row}"].value

            if category is not None:
                category = str(category).strip()

            if level_1 is not None:
                level_1 = str(level_1).strip()

            if level_2 is not None:
                level_2 = str(level_2).strip()

            if not category:
                continue

            if "total" in str(category).lower():
                continue

            if level_1 and "total" in str(level_1).lower():
                continue

            if level_2 and "total" in str(level_2).lower():
                continue

            if metric_1 in [0, 0.0, None, "-"]:
                continue

            key = make_lookup_key(category, level_1, level_2)
            lookup_result = lookup.get(key, {})

            if key not in lookup:
                log_entry["Lookup Misses"] += 1

            rows.append({
                "Period": log_entry["Period"],
                "Week": log_entry["Week"],
                "Source File": file_path.name,
                "Category": category,
                "Level 1": level_1,
                "Level 2": level_2,
                "Metric 1": metric_1,
                "Metric 2": lookup_result.get("Metric 2"),
                "Metric 3": lookup_result.get("Metric 3"),
                "Lookup Matched": key in lookup,
            })

        workbook.close()

        log_entry["Status"] = "Success"
        log_entry["Rows Extracted"] = len(rows)

        return rows, log_entry

    except Exception as error:
        log_entry["Status"] = "Error"
        log_entry["Reason"] = str(error)

        return rows, log_entry


def extract_all_files():
    all_rows = []
    log_rows = []

    files = get_weekly_files()

    print("Weekly files selected:", len(files))

    for file_path in files:
        print("Processing:", file_path.name)

        rows, log_entry = extract_file(file_path)

        all_rows.extend(rows)
        log_rows.append(log_entry)

    return pd.DataFrame(all_rows), pd.DataFrame(log_rows)


def build_validation_tables(dataframe):
    week_summary = (
        dataframe
        .groupby(["Week"], dropna=False)[["Metric 1", "Metric 2"]]
        .sum()
        .reset_index()
        .sort_values("Week")
    )

    level_summary = (
        dataframe
        .groupby(["Week", "Category", "Level 1", "Level 2"], dropna=False)[
            ["Metric 1", "Metric 2"]
        ]
        .sum()
        .reset_index()
        .sort_values(["Week", "Category", "Level 1", "Level 2"])
    )

    lookup_check = (
        dataframe
        .groupby(["Week", "Lookup Matched"], dropna=False)
        .size()
        .reset_index(name="Row Count")
    )

    return week_summary, level_summary, lookup_check


def export_results(dataframe, log_dataframe):
    week_summary, level_summary, lookup_check = build_validation_tables(dataframe)

    with pd.ExcelWriter(OUTPUT_FILE, engine="openpyxl") as writer:
        dataframe.to_excel(writer, sheet_name="Master_Output", index=False)
        log_dataframe.to_excel(writer, sheet_name="Extraction_Log", index=False)
        week_summary.to_excel(writer, sheet_name="Week_Summary", index=False)
        level_summary.to_excel(writer, sheet_name="Level_Summary", index=False)
        lookup_check.to_excel(writer, sheet_name="Lookup_Check", index=False)


if __name__ == "__main__":
    extracted_data, extraction_log = extract_all_files()
    export_results(extracted_data, extraction_log)

    print()
    print("Rows extracted:", len(extracted_data))
    print("Output created:")
    print(OUTPUT_FILE)
